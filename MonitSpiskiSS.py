#!/usr/bin/python3
# -*- coding: utf-8 -*-

# автор: Титовский С.А.
# 22/01/2025

import wx
from wx.adv import TaskBarIcon as TaskBarIcon
import sys
#import requests
#import re
import time
import string # for buttons
import datetime  # импорт библиотеки дат и времени
from datetime import datetime, timedelta
import os  # импорт библиотеки для работы с операционной системой
import socket
import sqlite3
import openpyxl # импорт библиотек для работы Excel
from openpyxl import Workbook # импорт библиотек для работы Excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.cell import get_column_letter
import locale # для работы с локалью
import threading
from threading import Thread # для работы с потоками
from pubsub import pub # для работы с подписчиками
import ftplib # для работы с ftp
#import queue
from HelloFrame import HelloFrame # import helloframe
from CheckFrame import ChFrame # import checkframe
from Logging import LogThread, ToLog # import logging
from SettingsDlg import SettingsDlg #import settingdlg
from RoitineFoo import DataForMChanges, DataOneMeeting # for parsing mchanges, pchanges

#=============================================
#=============================================
#=============================================
#=============================================
def except_foo_dec(foo):
    def wrapper(*args, **kwargs):
        #ToLog(f"{foo.__name__}, *args = {args}, **kwargs = {kwargs} started")
        try:
            return foo(*args, **kwargs)
        except Exception as Err:
            ToLog(f"Error in {foo.__name__}, *args = {args}, **kwargs = {kwargs}, Error code = {Err}")
            #raise Exception
        else:
            ToLog(f"{foo.__name__}, *args = {args}, **kwargs = {kwargs} finished successfully")
    return wrapper

def except_foo_brief(foo):
    def wrapper(*args, **kwargs):
        #ToLog(f"{foo.__name__} started")
        try:
            return foo(*args, **kwargs)
        except Exception as Err:
            ToLog(f"Error in {foo.__name__}, *args = {args}, **kwargs = {kwargs}, Error code = {Err}")
            #raise Exception
        else:
            ToLog(f"{foo.__name__} finished successfully")
    return wrapper

def info_foo_dec(foo):
    def wrapper(*args, **kwargs):
        #ToLog(f"{foo.__name__}, *args = {args}, **kwargs = {kwargs} started")
        return foo(*args, **kwargs)
        ToLog(f"{foo.__name__}, *args = {args}, **kwargs = {kwargs} finished")
    return wrapper

def except_method_dec(method):
    def wrapper(self, *args, **kwargs):
        #ToLog(f"{method.__name__}, *args = {args}, **kwargs = {kwargs} started")
        try:
            method(self, *args, **kwargs)
        except Exception as Err:
            ToLog(f"Error in {method.__name__}, *args = {args}, **kwargs = {kwargs}, Error code = {Err}")
            #raise Exception
        else:
            ToLog(f"{method.__name__}, *args = {args}, **kwargs = {kwargs} finished successfully")
    return wrapper

def except_method_brief(method):
    def wrapper(self, *args, **kwargs):
        #ToLog(f"{method.__name__} started")
        try:
            method(self, *args, **kwargs)
        except Exception as Err:
            ToLog(f"Error in {method.__name__}, *args = {args}, **kwargs = {kwargs}, Error code = {Err}")
            #raise Exception
        else:
            ToLog(f"{method.__name__} finished successfully")
    return wrapper

def except_method_briefer(method):
    def wrapper(self, *args, **kwargs):
        #ToLog(f"{method.__name__} started")
        try:
            method(self, *args, **kwargs)
        except Exception as Err:
            ToLog(f"Error in {method.__name__}, *args = {args}, **kwargs = {kwargs}, Error code = {Err}")
            #raise Exception
    return wrapper

def info_method_dec(method):
    def wrapper(self, *args, **kwargs):
        #ToLog(f"{method.__name__}, *args = {args}, **kwargs = {kwargs} started")
        method(self, *args, **kwargs)
        ToLog(f"{method.__name__}, *args = {args}, **kwargs = {kwargs} finished")
    return wrapper

#=================================
# HelloFrame        
class IconTray(TaskBarIcon):
    def __init__(self, frame, docdir = "somedir", stop = True, pub = "SomePub"):
        TaskBarIcon.__init__(self)
        self.frame = frame
        self.docdir = docdir
        self.stop = stop
        self.pub = pub
        #image
        self.SetIcon(wx.Icon(os.getcwd() + "\\images\\WritingIco.ico"), "MonitSpiskiSS")
        self.imageidx = 1
        #popup menu
        self.START_STOP = wx.NewIdRef(count = 1)
        self.CONVERT = wx.NewIdRef(count = 1)
        self.CH_FRAME = wx.NewIdRef(count = 1)
        self.SHOW_STNGS = wx.NewIdRef(count = 1)
        self.SHOW_LIC = wx.NewIdRef(count = 1)
        self.SHOW_VER = wx.NewIdRef(count = 1)
        self.EXIT_SCR = wx.NewIdRef(count = 1)
        #events
        #self.Bind(wx.adv.EVT_TASKBAR_LEFT_DCLICK, self.OnDClick)
        self.Bind(wx.EVT_MENU, self.OnStartStop, id = self.START_STOP)
        self.Bind(wx.EVT_MENU, self.OnConvert, id = self.CONVERT)
        self.Bind(wx.EVT_MENU, self.OnCheckFrame, id = self.CH_FRAME)
        self.Bind(wx.EVT_MENU, self.OnSettings, id = self.SHOW_STNGS)
        self.Bind(wx.EVT_MENU, self.OnLicense, id = self.SHOW_LIC)
        self.Bind(wx.EVT_MENU, self.OnVersion, id = self.SHOW_VER)
        self.Bind(wx.EVT_MENU, self.OnExit, id = self.EXIT_SCR)

    #@except_method_dec
    def CreatePopupMenu(self):
        menu = wx.Menu()
        if self.stop == True:
            menu.Append(self.START_STOP, "Start...")
        else:
            menu.Append(self.START_STOP, "Stop...")
        menu.Append(self.CONVERT, "Convert to Excel...")
        menu.Append(self.CH_FRAME, "Show status...")
        menu.Append(self.SHOW_STNGS, "Settings...")
        menu.Append(self.SHOW_LIC, "License")
        menu.Append(self.SHOW_VER, "Version Info")
        menu.Append(self.EXIT_SCR, "Exit")
        return menu

    @except_method_dec
    def OnStartStop(self, evt):
        wx.CallAfter(
            pub.sendMessage, self.pub,
            mess = "StartStop")

    @except_method_brief
    def OnConvert(self, evt):
        if self.docdir + "\\temp\\temp.db" in os.listdir(self.docdir + "\\temp"):
            os.remove(self.DocDir + "\\temp\\temp.db")
            ToLog("Previous temp file removed")
                                                              
        DialogLoad = wx.FileDialog(
            None,
            "Load file",
            #defaultDir = self.DocDir,
            wildcard = "db files (*.db)|*db",
            style = wx.FD_OPEN)

        if DialogLoad.ShowModal() == wx.ID_CANCEL:
            ToLog("Cancel converting")
            return

        else:     
            #print("GetDir = ", DialogLoad.GetDirectory())
            #print("GetFilename = ", DialogLoad.GetFilename())
            selected_file = DialogLoad.GetDirectory() + "\\" + DialogLoad.GetFilename()
            selected_dir = DialogLoad.GetDirectory()
            
            tempFile = self.docdir + "\\temp\\temp.db"
            CopyFile(selected_file, tempFile)
            ToLog(f"Converting f{selected_file} to result.xlsx")
            
            ConvThread = ConvertThread(path_to_file = tempFile, path_to_dir = selected_dir)
            ConvThread.setDaemon(True)
            ConvThread.start()

    @except_method_dec
    def OnCheckFrame(self, evt):
        wx.CallAfter(
            pub.sendMessage, self.pub,
            mess = "CheckFrame")
    
    @except_method_dec
    def OnSettings(self, evt):
        wx.CallAfter(
            pub.sendMessage, self.pub,
            mess = "Settings")

    @except_method_brief
    def OnLicense(self, evt):
        LICENSE = (
            "Данная программа является свободным программным обеспечением\n"+
            "Вы вправе распространять её и/или модифицировать в соответствии\n"+
            "с условиями версии 2 либо по Вашему выбору с условиями более\n"+
            "поздней версии Стандартной общественной лицензии GNU, \n"+
            "опубликованной Free Software Foundation.\n\n\n"+
            "Эта программа создана в надежде, что будет Вам полезной, однако\n"+
            "на неё нет НИКАКИХ гарантий, в том числе гарантии товарного\n"+
            "состояния при продаже и пригодности для использования в\n"+
            "конкретных целях.\n"+
            "Для получения более подробной информации ознакомьтесь со \n"+
            "Стандартной Общественной Лицензией GNU.\n\n"+
            "Данная программа написана на Python\n"
            "Автор: Титовский С.А.")
        
        wx.MessageBox(LICENSE, "Лицензия", wx.OK)

    @except_method_brief
    def OnVersion(self, evt):
        txt = (
            "\n\nVersion of Script 22.01.2025" +
            "\n\nAdded: " +
            "\n\t- Function of checking meeting participants (store it in several tables of mchanges.db)" +
            "\n\t- AutoCleaning old Logs, on FTP and on Local"
            "\n\nVersion of Script 25.12.2024" +
            "\n\nAdded: " +
            "\n\t- Question of start monitoring when executed" +
            "\n\t- Field of current start/stop status" + 
            "\n\nModified: " +
            "\n\t- Flag of start, pause checking in mchanges.db\n")
        
        wx.MessageBox(txt, "Version Info", wx.OK)

    @except_method_dec
    def OnExit(self, evt):
        wx.CallAfter(
            pub.sendMessage, self.pub,
            mess = "ExitCmd")
        
#=========================================
#=========================================
#=========================================
#=========================================        
# OsnFrame
class Main_Frame(wx.Frame):

    def __init__(self, parent, WinPos = wx.DefaultPosition, DateUser = "01.01.2021", DocDir = os.getcwd()):
        wx.Frame.__init__(
            self, parent, -1, "Список совещаний")

        self.WinPos = WinPos
        self.Date = Dt_to_txt(datetime.today())
        self.DocDir = DocDir

        frameIcon = wx.Icon(os.getcwd() + "\\images\\WritingPNG.png")
        self.SetIcon(frameIcon)

        #self.pub = f"Main_Frame {self.GetId()}"
        self.pub = "Main_Frame"
        pub.subscribe(self.UpdateDisplay, self.pub)
        
        self.OpenPanel()

#=========================================
    @info_method_dec
    def OpenPanel(self):
        self.panel = Main_Panel(self, self.Date, self.DocDir, self.pub)

#==========================================
    #@except_method_dec
    def UpdateDisplay(self, mess = "someMessage"):
        #print("\t\tGOT mess = " + str(mess))
        if isinstance (mess, str):
            #print("\t\tGOT mess = " + mess)
            if mess == "RenewNow":
                self.panel.RenewCommand()
            if mess == "RenewNowFirst":
                self.panel.RenewCommand(first_time = True)
            elif mess == "ExitCmd":
                self.panel.CloseCmd(from_tray = True)
            elif mess == "Settings":
                self.panel.ShowSettings()
            elif mess == "StartStop":
                self.panel.StartStop()
            elif mess == "CheckFrame":
                self.panel.CheckFrame()
        if isinstance (mess, list):
            #print("\t\tGOT mess = " + mess[0])
            if mess[0] == "ListDate":
                self.panel.CheckDate(first_time = mess[1], date = mess[2], data = mess[3])
            elif mess[0] == "ChangedDate":
                self.panel.ChangeDate(mess[1])
            elif mess[0] == "FTPtime":
                self.panel.chData.update({"Last FTP sending time": mess[1]})
            elif mess[0] == "AddToPChanges":
                self.panel.ToPChanges(
                    date = mess[1], someid = mess[2],
                    listdata = mess[3], comment = mess[4])
                    
#=========================================
#=========================================
#=========================================
#=========================================
# panel
class Main_Panel(wx.Panel):
    def __init__(self, parent, date, docdir, pub):
        wx.Panel.__init__(self, parent = parent)

        self.frame = parent
        self.DocDir = docdir
        self.date = date
        self.pub = pub

        self.settings = {}
        self.LoadSettings()

        #change here num of days
        self.threads = None
        self.DThreads = {}
        self.MThreads = {}
        self.actualData = None
        self.conns = {}
        #self.mconns = {}
        self.PQueueItems = {}
        self.pause = True
        self.RenewThread = None
        self.RenewEvt = threading.Event()

        #chframe
        self.chFrame = None
        self.chData = {
            "Current threads": "not updated",
            "Last databases renew time": "not updated",
            "Last FTP sending time": "not updated",
            "Script started at time": str(datetime.today()),
            "This frame opened at time": "not updated",
            "Scan Status": "Paused"}
        self.FTPtime = None
        
        self.SetSize((300, 400))
        self.frame.Bind(wx.EVT_CLOSE, self.OnCloseWindow)
        self.Show(True)

        self.Tray = IconTray(self.frame, docdir = self.DocDir, stop = self.pause, pub = self.pub)
        self.CheckFrame()
        self.AskForStart()
        
#=============================================
    @except_method_brief
    def AskForStart(self):
        dlg = wx.MessageDialog(
            self, "Do you want to start checking meetings?",
            " ", wx.YES_NO)
        answer = dlg.ShowModal()
        
        if answer == wx.ID_YES:
            print("Pressed start")
            ToLog("Pressed Start Checking")
            self.StartStop()

        elif answer == wx.ID_NO:
            ToLog("Pressed No Start Checking")

    except_method_brief
    def LoadSettings(self):
        if "settings.db" not in os.listdir(self.DocDir + "\\Based"):
            #print("Create new db file")
            self.CreateSettingsDB(path = self.DocDir + "\\Based\\settings.db")
            
        self.LoadFrDB(path = self.DocDir + "\\Based\\settings.db")
        
        ToLog("Settings after load = " + str(self.settings))

    @except_method_dec
    def LoadFrDB(self, path = "somepath"):
        setconn = sqlite3.connect(path)
        cursor = setconn.execute("SELECT * FROM SETTINGS WHERE show_status=(?)", (1, ))
        pair = [(row[1], row[2]) for row in cursor]
        setconn.close()

        for i in range (0, len(pair)):
            self.settings.update({pair[i][0]:pair[i][1]})

        return 

    @except_method_dec
    def CreateSettingsDB(self, path = "somepath"):
        setconn = sqlite3.connect(path)
        setconn.execute(
            "CREATE TABLE IF NOT EXISTS SETTINGS" +
            "(num INTEGER PRIMARY KEY AUTOINCREMENT," +
            "key TEXT NOT NULL," +
            "value TEXT NOT NULL," +
            "show_status BOOLEAN NOT NULL CHECK(show_status IN (0, 1))," + 
            "note1 CHAR(100)," +
            "note2 CHAR(100));")

        dictDef = {
            "num_days": "4", "time_renew": "60", "time_ftp": "1200",
            "ftp_addr": "10.135.11.177", "ftp_login": "login",
            "ftp_password": "TSA&44186"}

        for item in dictDef.keys():
            setconn.execute(
                "INSERT INTO SETTINGS VALUES (" +
                "NULL, '" + item + "', '" + dictDef[item] + "', 1, NULL, NULL)")

        setconn.commit()
        setconn.close()

    @except_method_dec
    def SaveSettings(self):
        if "settings.db" not in os.listdir(self.DocDir + "\\Based"):
            self.CreateSettingsDB(path = self.DocDir + "\\Based\\settings.db")
        else:
            self.SaveDB(path = self.DocDir + "\\Based\\settings.db")

    @except_method_dec
    def SaveDB(self, path = "somepath"):
        setconn = sqlite3.connect(path)
        for item in self.settings.keys():
            ToLog(f"Updated values in settings db =  {item}: {self.settings[item]}")
            cursor = setconn.execute("UPDATE SETTINGS SET value = '" + self.settings[item] + "' WHERE key = '" + item + "'")
        setconn.commit()
        setconn.close()

    @except_method_brief
    def ShowSettings(self):
        dlg1 = SettingsDlg(label = "Settings of MonitSpiskiSS", settings = self.settings)
        if dlg1.ShowModal() == wx.ID_OK:
            temp = {}
            temp.update({"num_days": dlg1.Data[0].GetStringSelection()})
            temp.update({"time_renew": dlg1.Data[1].GetStringSelection()})
            temp.update({"time_ftp": dlg1.Data[2].GetStringSelection()})
            temp.update({"ftp_addr": dlg1.Data[3].GetValue()})
            temp.update({"ftp_login": dlg1.Data[4].GetValue()})
            temp.update({"ftp_password": dlg1.Data[5].GetValue()})

        if temp == self.settings:
            ToLog("After SettingsDlg no changes entered")
            return

        self.settings = temp
        self.SaveSettings()
        
        ToLog("Stopping threads after changing settings")
        self.StopThreads(except_renew = True)
        ToLog("Starting threads with new settings")
        self.StartScanDays(days = int(self.settings["num_days"]))
        if self.RenewThread:
            self.RenewThread.UpdateData(upd_data = self.settings)
        
    @except_method_brief
    def CheckFrame(self):
        if self.chFrame != None:
            try:
                ToLog("Destroy previous CheckFrame")
                self.chFrame.Destroy()
            except Exception:
                pass
            self.chFrame = None
        self.chData.update({"This frame opened at time": str(datetime.today())})
        self.chFrame = ChFrame(
            label = "Check status", data = self.chData,
            path_to_png = os.getcwd() + "\\images\\WritingPNG.png")                    
 
    @except_method_dec  
    def OnCloseWindow(self, evt):
        self.CloseCmd()
        evt.Skip()
        sys.exit()

    @except_method_dec 
    def CloseCmd(self, from_tray = False):
        ToLog("Application closed by User's command")
        self.Show(False)
        ToLog(f"Save and close db file, now conns = {self.conns}")
        #print(str(self.conns.items()))
        self.AddToDB(mess = "Application closed")
        for key in self.conns.keys():
            self.conns[key].commit()
            self.conns[key].close()
            
        self.StopThreads(stop_MThreads = True)
       
        ToLog("Stopping LogThread")
        global threadLog
        threadLog.stop = True
        threadLog.join()

        if from_tray == True:
            self.frame.Destroy()
            sys.exit()
            
    @except_method_dec 
    def StartRenewThread(self):
        self.RenewThread = RenewThread(evt = self.RenewEvt, pub = self.pub, settingsDict = self.settings, testChDate = False)
        self.RenewThread.setDaemon(True)
        self.RenewThread.start()

    @except_method_dec 
    def StopThreads(self, except_renew = False, stop_MThreads = False):
        if except_renew == False:
            if self.RenewThread:
                self.RenewThread.Stop()

        templist = list(self.DThreads.keys())[:]
        for key in templist:
            if self.DThreads[key].is_alive() == True:
                self.DThreads[key].Stop()
            del self.DThreads[key]
        if stop_MThreads == True:
            templist = list(self.MThreads.keys())[:]
            for key in templist:
                if self.MThreads[key].is_alive() == True:
                    self.MThreads[key].Stop()
                del self.MThreads[key]
                
  
    @except_method_dec
    def StartStop(self):
        if self.RenewEvt.isSet():
            #print("Pause Renew")
            self.Tray.stop = True
            self.RenewEvt.clear()
            self.AddToDB(mess = "Application paused")
            self.chData.update({"Scan Status": "Paused"})
        else:
            if self.RenewThread:
                print("Resume")
                self.Tray.stop = False
                self.RenewEvt.set()
                self.AddToDB(mess = "Application resumed")
                self.chData.update({"Scan Status": "Running"})
            else:
                print("Start new thrings")
                self.Tray.stop = False
                self.RenewEvt.set()
                self.StartScanDays(days = int(self.settings["num_days"]))
                self.StartRenewThread()
                self.chData.update({"Scan Status": "Running"})
        if self.chFrame:
            self.chFrame.UpdateData(upd_data = self.chData) 

    @except_method_dec
    def AddToDB(self, mess = "None"):
        if self.conns:
            for key in self.conns.keys():
                #print(f"add to db {key} message {mess}")
                self.conns[key].execute(
                    "INSERT INTO MEETING_CHANGES VALUES (" +
                    "NULL,'"  + Dt_to_txt(datetime.now()) + "','" +
                    str(datetime.now())[11:16] + "','None'," +
                    "'None', 'None', 'None', 'None', 'None', 'None', 'None', '" + mess + "', NULL)")
                self.conns[key].commit()
            
    @except_method_dec
    def StartScanDays(self, days = 4):
        self.actualData = {}
        for day in range (0, days):
            date_tm = datetime.today() + timedelta(days = day)
            date = Dt_to_txt(date_tm)
    
            thread = ScanDayThread(pub = self.pub, date = date)
            thread.setDaemon(True)
            thread.start()
            self.DThreads.update({date: thread})
        print("now keys of DTHREADS = " + str(self.DThreads.keys()))
        self.chData.update({
            "Current threads": str(days) + " days watching: " + ", ".join(list(self.DThreads.keys()))})

        if self.chFrame:
            self.chFrame.UpdateData(upd_data = self.chData)

    @except_method_briefer
    def RenewCommand(self, first_time = False):
        if first_time == False:
            self.chData.update({"Last databases renew time": str(datetime.today())})
        if self.FTPtime:
            self.chData.update({"Last FTP sending time": self.FTPtime})                   
        templist = []
        for key in self.DThreads.keys():
            self.DThreads[key].ScanNow(first_time)
            templist.append(key)
        
        self.chData.update({
            "Current threads": str(len(templist)) + " days watching: " + ", ".join(templist)})

        if self.chFrame:
            self.chFrame.UpdateData(upd_data = self.chData)

    @except_method_briefer
    def CheckDate(self, first_time, date, data):
        if first_time == False:
            if datetime.strptime(date, "%d.%m.%Y") < (datetime.strptime(Dt_to_txt(datetime.today()), "%d.%m.%Y")):
                print(f"Received data from past days {date}, deleting this date from threads")
                ToLog(f"Received data from past days {date}, deleting this date from threads")
                self.PopOldDate(date = date, source = "CheckDate foo")
            else:
                self.RenewData(first_time, date, data)  
        else:
            self.RenewData(first_time, date, data)
        
    @except_method_dec
    def ChangeDate(self, date):     
        text_date = Dt_to_txt(date)
        keys = list(self.DThreads.keys())[:]
        survived = []
        
        for key in keys:
            temp_date = self.DThreads[key].date
            #if (datetime.strptime(thread.date, "%d.%m.%Y") + timedelta(days = 1)).date() < date.date():
            if datetime.strptime(temp_date, "%d.%m.%Y").date() < date.date():
                ToLog(f"Deleting ScanThread with old date {temp_date}")
                print(f"Deleting ScanThread with old date {temp_date}")
                self.PopOldDate(date = temp_date, source = "ChangeDate")
               
                #self.CleanOldLogFiles(keys = ["mchanges_"])
            else:
                survived.append(temp_date)
        
        new_data_list = []
        for day in range (0, int(self.settings["num_days"])):
            temp_day = date + timedelta(days = day)
            new_data_list.append(Dt_to_txt(temp_day))
        print("new_data_list = " + str(new_data_list))
        print("survived = " + str(survived))
            
        new_datas = list(set(new_data_list) - set(survived))[:]
        print("New Datas = " + str(new_datas))
        if len(new_datas) > 0:
            for day in new_datas:
                thread = ScanDayThread(pub = self.pub, date = day)
                thread.setDaemon(True)
                thread.start()
                self.DThreads.update({day: thread})
                self.CreateNewDB(day)
        print("now keys of DTHREADS = " + str(self.DThreads.keys()))
        self.chData.update({
            "Current threads": str(len(self.DThreads.keys())) + " days watching: " + ", ".join(list(self.DThreads.keys()))})

        if self.chFrame:
            self.chFrame.UpdateData(upd_data = self.chData)

    @except_method_dec
    def PopOldDate(self, date = "01.01.2012", source = "some_source", except_DThreads = False):
        global LogDir
        global MonitLogDir
        print("from source = " + source)
        templist = list(self.MThreads.keys())[:]
        if date in templist:
            if self.MThreads[date].is_alive() == True:
                self.MThreads[date].Stop()
            del self.MThreads[date]
        if except_DThreads == False:
            templist = list(self.DThreads.keys())[:]
            if date in templist:
                if self.DThreads[date].is_alive() == True:
                    self.DThreads[date].Stop()
                del self.DThreads[date]
        templist = list(self.actualData.keys())[:]
        if date in templist:
            del self.actualData[date]
        templist = list(self.conns.keys())[:]
        if date in templist:
            self.conns[date].commit()
            self.conns[date].close()
            del self.conns[date]
        templist = list(self.PQueueItems.keys())[:]
        if date in templist:
            del self.PQueueItems[date]

        ToLog("now connkeys = " + str(self.conns.keys()))
        ToLog("now DThreadkeys = " + str(self.DThreads.keys()))
        ToLog("now MThreadkeys = " + str(self.MThreads.keys()))
        ToLog("now act datakeys = " + str(self.actualData.keys()))
        ToLog("now PQ datakeys = " + str(self.PQueueItems.keys()))
        
        print("now connkeys = " + str(self.conns.keys()))
        print("now DThreadkeys = " + str(self.DThreads.keys()))
        print("now MThreadkeys = " + str(self.MThreads.keys()))
        print("now act datakeys = " + str(self.actualData.keys()))
        print("now PQ datakeys = " + str(self.PQueueItems.keys()))

        ClearLogs(LogDir)
        ClearLogs(MonitLogDir)
                
    @except_method_briefer
    def RenewData(self, first_time, date, data):
        self.AddToPQueue(date, data.keys())
        if len(self.actualData) == 0:
            self.actualData.update({date: data})
            self.SaveToDB(first_time = first_time, date = date, data = data)
        elif date in self.actualData:
            self.CompareData(date = date, data = data)
        else:
            self.actualData.update({date: data})
            self.SaveToDB(first_time = first_time, date = date, data = data)
        #print(f"Now ActualData\n {self.actualData}")

        if self.chFrame:
            self.chFrame.UpdateData(upd_data = self.chData)
        if len(list(self.DThreads.keys())) < len(list(self.conns.keys())):
            templist = list(self.conns.keys())[:]
            for item in templist:
                if item not in list(self.DThreads.keys()):
                    self.PopOldDate(date = item, source = "RenewData", except_DThreads = True)

    @except_method_briefer
    def CompareData(self, date, data):
        if self.actualData[date] == data:
            #ToLog(f"No changes in actualData on date {date}")
            pass
        else:
            for item in data.keys():
                #if new meeting added
                if item not in self.actualData[date].keys():
                    ToLog("Meeting added:\n\t " + str(data[item]))
                    self.actualData[date].update({item: data[item]})
                    self.SaveToDB(first_time = False, date = date, item = data[item])
                else:
                    if data[item] == self.actualData[date][item]:
                        #ToLog(f"No changes in item {item}")
                        pass
                    else:
                        ToLog(
                            f"Changes in item {item}:\n\t previous field = " +
                            f"{self.actualData[date][item]}, \n\t new field = {data[item]}")
                        self.SaveToDB(first_time = False, date = date, item = data[item])
            #if meeting popped
            diff = list(set(self.actualData[date].keys()) - set(data.keys())) 
            if len(diff) > 0:
                self.PopMeetings(date, diff)
            self.actualData.update({date: data})

    @except_method_briefer
    def CreateNewDB(self, date = "01.01.2024"):
        nameDB = self.DocDir + "\\Monitoring_Logs\\mchanges_" + date + ".db"
        if "mchanges_" + date + ".db" in os.listdir(self.DocDir + "\\Monitoring_Logs"):
            if date not in self.conns.keys():
            #ToLog(f"mchanges_{date}.db is already in dir {self.DocDir}\\Monitoring_Logs")
                conn = sqlite3.connect(nameDB)
                self.conns.update({date: conn})
            
        else:
            #print(f"cerate new db with date {text_date}")
            ToLog(f"creating mchanges_{date}.db in dir {self.DocDir}\\Monitoring_Logs")
            conn = sqlite3.connect(nameDB)
            self.conns.update({date: conn})
            conn.execute(
                "CREATE TABLE IF NOT EXISTS MEETING_CHANGES" +
                "(num INTEGER PRIMARY KEY AUTOINCREMENT," +
                "date_now TEXT NOT NULL," +
                "time_now TEXT NOT NULL," +
                "date TEXT NOT NULL," +
                "id TEXT NOT NULL," +
                "studia TEXT NOT NULL," +
                "initiator TEXT NOT NULL," +
                "time TEXT NOT NULL," +
                "rezhim TEXT NOT NULL," +
                "theme TEXT," +
                "participants TEXT," + 
                #"show_status BOOLEAN NOT NULL CHECK(show_status IN (0, 1)),"
                "note1 CHAR(100)," +
                "note2 CHAR(100));")
            self.conns.update({date: conn})

    @except_method_briefer
    def SaveToDB(self, first_time = False, date = "01.01.2024", data = None, item = None):
        self.CreateNewDB(date = date)
        #print(f"I received data = {data}, item = {item}")
        if item:
            if first_time == False:
                self.conns[date].execute(
                    "INSERT INTO MEETING_CHANGES VALUES (" +
                    "NULL,'"  + Dt_to_txt(datetime.now()) + "','" +
                    str(datetime.now())[11:16] + "','" + date + "','" +
                    "','".join(item) + "',NULL,NULL)")
            else:
                self.conns[date].execute(
                    "INSERT INTO MEETING_CHANGES VALUES (" +
                    "NULL,'"  + Dt_to_txt(datetime.now()) + "','" +
                    str(datetime.now())[11:16] + "','" + date + "','" +
                    "','".join(item) + "','added at start',NULL)")      
            self.conns[date].commit()

        if data:       
            for key in data.keys():
                self.SaveToDB(first_time = first_time, date = date, item = data[key])
                
    @except_method_dec
    def PopMeetings(self, date = "01.01.2024", id_list = ["111111"]):
        print("popping " + ", ".join(id_list))
        for item in id_list:
            self.conns[date].execute(
                "INSERT INTO MEETING_CHANGES VALUES (" +
                "NULL,'"  + Dt_to_txt(datetime.now()) + "','" +
                str(datetime.now())[11:16] + "','" + date + "','" +
                item + "','Deleted','Deleted','Deleted','Deleted','Deleted','Deleted',NULL,NULL)")
            self.conns[date].commit()

    #@except_method_dec
    def AddToPQueue(self, date = "somedate", keys = "somekeys"):
        if date not in self.PQueueItems.keys():
            self.PQueueItems.update({date: set(keys)})
            self.CheckMThreads(date = date, newid = True)
        else:
            if set(keys) == self.PQueueItems[date]:
                #print(f"no changes in ids list on date {date}")
                self.CheckMThreads(date = date, newid = False)
            else:
                self.PQueueItems.update({date: set(keys)})
                self.CheckMThreads(date = date, newid = True)

    @except_method_briefer
    def CheckMThreads(self, date = "somedate", newid = True):
        if date in self.MThreads.keys():
            if self.MThreads[date].is_alive() == True:
                #print(f"Mthread {date} is alive, kick him")
                self.MThreads[date].Kick()
                if newid == True:
                    #print(f"Mthread {date} is alive, update data to " + str(self.PQueueItems[date]))
                    self.MThreads[date].newids = self.PQueueItems[date]
                    
            else:
                #print(f"Mthread {date} not alive, starting new")
                thread = MeetingThread(date = date, ids = self.PQueueItems[date], docdir = self.DocDir, pub = self.pub)
                thread.setDaemon(True)
                thread.start()
                self.MThreads.update({date: thread})
        else:
            #print(f"No such Mthread {date}, creating new")
            thread = MeetingThread(date = date, ids = self.PQueueItems[date], docdir = self.DocDir, pub = self.pub)
            thread.setDaemon(True)
            thread.start()
            self.MThreads.update({date: thread})

    @except_method_briefer
    def ToPChanges(self, date = "somedate", someid = "someid", listdata = [], comment = None):

        #if no such date in conns
        if date not in self.DThreads.keys():
            ToLog(f"Ignoring ToPChanges from {date}, it is not in PThread keys:" + str(self.DThreads.keys()))
            print(f"Ignoring ToPChanges from {date}, it is not in PThread keys:" + str(self.DThreads.keys()))
            return
            #nameDB = self.DocDir + "\\Monitoring_Logs\\mchanges_" + date + ".db"
            #self.conns.update({date: sqlite3.connect(nameDB)})

        #create table if not exist
        tablename = "id" + someid
        self.conns[date].execute(
            "CREATE TABLE IF NOT EXISTS '" + tablename + "'" + 
            "(num INTEGER PRIMARY KEY AUTOINCREMENT," +
            "date_now TEXT NOT NULL," +
            "time_now TEXT NOT NULL," +
            "id TEXT NOT NULL," +
            "region TEXT NOT NULL," +
            "room TEXT NOT NULL," +
            "dolgnost TEXT," +
            "fio TEXT," +
            "note TEXT," + 
            #"show_status BOOLEAN NOT NULL CHECK(show_status IN (0, 1)),"
            "note1 CHAR(100)," +
            "note2 CHAR(100));")

        #add listdata to table
        for item in listdata:
            self.conns[date].execute(
                "INSERT INTO '" + tablename + "' VALUES (" +
                "NULL,'"  + Dt_to_txt(datetime.now()) + "','" +
                str(datetime.now())[11:16] + "','" + someid + "','" + 
                "','".join(item) + "','" +
                str(comment) + "',NULL)")

        self.conns[date].commit()  
        print(f"{comment} some to pchanges_{date}.db")
            
    @except_method_brief
    def OnBtn(self, evt):
        ToLog("OnBtn pressed")

#=============================================
#=============================================
#=============================================
#=============================================
#ScanDayThread
class ScanDayThread(threading.Thread):
    def __init__(self, num_day = 0, pub = "my_pub", date = None):
        super().__init__()
        self.stop = False
        self.pub = pub
        if date:
            self.date = str(date)
        else:
            date_tm = datetime.today() + timedelta(days = num_day)
            self.date = Dt_to_txt(date_tm)

    @except_method_brief
    def run(self):
        ToLog(f"Thread on date {self.date} started")
        while True:
            if self.stop == True:
                break
            time.sleep(1)
        ToLog(f"Thread on date {self.date} finished")

    @except_method_briefer
    def ScanNow(self, first_time = True):
        ListDate = DataForMChanges(date = self.date)
        wx.CallAfter(
            pub.sendMessage, self.pub,
            mess = ["ListDate", first_time, self.date, ListDate])
        
    @except_method_brief        
    def Stop(self):
        ToLog(f"Thread on date {self.date} received Stop command")
        self.stop = True

#=============================================
#=============================================
#=============================================
#=============================================
#ScanMeetingThread
class MeetingThread(threading.Thread):
    def __init__(self, date = "01.01.2024", ids = {"000000"}, docdir = "somedir", pub = "somepub"):
        super().__init__()
        self.stop = False
        self.pub = pub
        self.date = date
        self.ids = ids
        self.newids = ids
        self.docdir = docdir
        self.checking = True
        self.checknow = True
        self.dictData = {}
        
    @except_method_brief
    def run(self):
        ToLog(f"Meeting Thread on date {self.date} started")
        while True:
            if self.stop == True:
                break
            if self.checknow == False:
                time.sleep(2)
                continue
            self.ParseData(self.ids)
            self.checknow = False
            self.ids = self.newids

        ToLog(f"Meeting Thread on date {self.date} finished")

    @except_method_briefer
    def Kick(self):
        #print(f"kicked mthread {self.date}")
        self.checknow = True

    @except_method_briefer
    def ParseData(self, ids):
        for someid in ids:
            if self.stop == True:
                break
            DataId = DataOneMeeting(idsov = someid)
            self.CompareData(someid, DataId)

    @except_method_briefer
    def CompareData(self, someid, DataId):
        if someid not in self.dictData.keys():
            self.dictData.update({someid: DataId})
            self.WriteToConn(someid = someid, listdata = DataId, comment = "added first time")
        else:
            if self.dictData[someid] == DataId:
                return
            else:
                #check difference detween sets
                try:
                    if set(self.dictData[someid]) != set(DataId):
                        self.WriteToConn(
                            someid = someid,
                            listdata = list(set(self.dictData[someid]) - set(DataId)),
                            comment = "deleted")
                        self.WriteToConn(
                            someid = someid,
                            listdata = list(set(DataId) - set(self.dictData[someid])),
                            comment = "added")
                    self.dictData.update({someid: DataId})
                except Exception:
                    print("ERROR")
                        
    @except_method_brief        
    def Stop(self):
        ToLog(f"Meeting Thread on date {self.date} received Stop command")
        self.stop = True

    @except_method_briefer
    def WriteToConn(self, someid = "someid", listdata = [], comment = "default comment"):
        if len(listdata) == 0:
            return
        wx.CallAfter(
            pub.sendMessage, self.pub,
            mess = ["AddToPChanges", self.date, someid, listdata, comment])

#=============================================
#=============================================
#=============================================
#=============================================
#RenewThread
class RenewThread(threading.Thread):
    @except_method_dec
    def __init__(self, evt, pub = "somePub", settingsDict = {"some": "some"}, testChDate = False):
        super().__init__()
        self.stop = False
        self.evt = evt
        self.pub = pub
        self.settings = settingsDict
        self.TimeRenew = int(self.settings["time_renew"])
        self.TimeFTP = int(self.settings["time_ftp"])
        self.cycles = self.TimeFTP // self.TimeRenew
        self.today = datetime.today().date()
        self.testChDate = testChDate
        #print("at start today = " + str(self.today))
        #self.evt.set()
        self.once = 0
        
    @except_method_brief
    def run(self):
        #print(f"Renew thread started wuth pub {self.pub} and timerenew {self.TimeRenew}")
        ToLog(f"Renew thread started wuth pub {self.pub} and timerenew {self.TimeRenew}")
        thread = FTPThread(self.settings, self.pub)
        thread.setDaemon(True)
        thread.start()
        self.RenewThreadCommand(first_time = True)
        startTime = time.time()
        now_cycles = 0
        if self.testChDate == True:
            offset = 1
        
        while True:
            #print("--iter--")
            self.evt.wait()
            
            if self.stop == True:
                ToLog("RenewThread stopped by Stop event")
                break
            if time.time() - startTime < self.TimeRenew:
                #ToLog("\t\tToo early for renew, last time = " + str(time.time() - startTime))
                time.sleep(2)
                continue
            else:
                self.RenewThreadCommand()
                if self.testChDate == True:
                    if now_cycles % 2 == 0:
                        self.ChangedDate(date = (datetime.today() + timedelta(days = offset)))
                        print("temporary incrasing ate by 1")
                        self.once = 1
                        offset = offset + 1

                else:
                    if datetime.today().date() != self.today:
                        self.ChangedDate(date = datetime.today())

                startTime = time.time()
                now_cycles += 1
                if now_cycles >= self.cycles:
                    thread = FTPThread(self.settings, self.pub)
                    thread.setDaemon(True)
                    thread.start()
                    now_cycles = 0
                    
            #ToLog("RenewThread Finished with RenewTime = " + str(TimeRenew))

    @except_method_brief
    def Stop(self):
        ToLog(f"Renew thread received Stop command")
        self.stop = True

    @except_method_briefer
    def RenewThreadCommand(self, first_time = False):
        if first_time == False:
            ToLog(f"Threads will be renewed because of time Renew = {self.TimeRenew} elapsed")
            wx.CallAfter(pub.sendMessage, self.pub, mess = "RenewNow")
        else:
            ToLog(f"Threads will be renewed because it's just started")
            wx.CallAfter(pub.sendMessage, self.pub, mess = "RenewNowFirst")

    @except_method_dec
    def ChangedDate(self, date = datetime.today().date()):
        ToLog(f"Changed date to {date}, sending ChangedDate command")
        print(f"Changed date to {date}, sending ChangedDate command")
        self.today = date.date()
        wx.CallAfter(pub.sendMessage, self.pub, mess = ["ChangedDate", date])

    @except_method_briefer
    def UpdateData(self, upd_data = {"somedata": "data"}):
        #print("upd_data")
        self.settings = upd_data
        self.TimeRenew = int(self.settings["time_renew"])
        self.TimeFTP = int(self.settings["time_ftp"])
        self.cycles = self.TimeFTP // self.TimeRenew
        self.evt.set()
        thread = FTPThread(self.settings, self.pub)
        thread.setDaemon(True)
        thread.start()

#=============================================
#=============================================
#=============================================
#=============================================
#RenewThread
class FTPThread(threading.Thread):
    @except_method_brief
    def __init__(self, settingsDict = {"some": "some"}, pub = "somepub"):
        super().__init__()
        self.settings = settingsDict
        self.pub = pub

    @except_method_dec
    def run(self):
        #print(f"FTP thread started")
        ToLog(f"FTP thread started")
        self.SendToFTP(path = self.settings["ftp_addr"], login = self.settings["ftp_login"], password = self.settings["ftp_password"], keys = ["mchanges_"])
        self.CleanFTP(
            path = self.settings["ftp_addr"], login = self.settings["ftp_login"], password = self.settings["ftp_password"], keys = ["mchanges_"],
            past_level = int(-15), future_level = int(10))
        #print(f"FTP thread finished")
        ToLog(f"FTP thread finished")

    def SendToFTP(self, path = "10.135.11.177", login = "DB", password = "TSA&44186", keys = ["somekey"]):
        global MonitLogDir
        try:
            namedir = str(socket.gethostname()).replace("\\", " ")
            namedir = namedir.replace(":", " ")
            ftp = ftplib.FTP(path)
            ftp.login(login, password)
            self.CreateDir(ftp, "MonitSpiskiSS\\" + namedir)

            for key in keys:
                for file in os.listdir(MonitLogDir):
                    if file.find(key) != -1 and file.find(".db") != -1:
                        with open(MonitLogDir + "\\" + file, "rb") as somefile:
                            ftp.storbinary("STOR " + file, somefile)
                            somefile.close()
                            ToLog(f"send to FTP file {file}")
                    else:
                        ToLog(f"not send to FTP {file}, it is not auto database file with key {key}")
                        #print(f"not send to FTP {file}, it is not auto database file with key {key}")
                    
            ftp.quit()
            #wx.CallAfter(pub.sendMessage, self.pub, mess = ["FTPtime", str(datetime.today())])
        except Exception as Err:
            print("Error connecting to FTP")
            ToLog("Error connecting to FTP, Error code = " + str(Err))

    def CleanFTP(
        self, path = "10.135.11.177", login = "DB", password = "TSA&44186", keys = ["somekey"],
        past_level = int(-20), future_level = int(20)):
        try:
            namedir = str(socket.gethostname()).replace("\\", " ")
            namedir = namedir.replace(":", " ")
            ftp = ftplib.FTP(path)
            ftp.login(login, password)
            self.CreateDir(ftp, "MonitSpiskiSS\\" + namedir)

            datelist = []
            for day in range (past_level, future_level):
                date_tm = datetime.today() + timedelta(days = day)
                datelist.append("mchanges_" + Dt_to_txt(date_tm) + ".db")
            #print("datelist for save = " + str(datelist))

            for item in ftp.nlst():
                if item not in datelist:
                    ftp.delete(item)
                    ToLog(f"From FTP deleted file {item}")
                    #print(f"I want delete {item}")
                
            ftp.quit()
            wx.CallAfter(pub.sendMessage, self.pub, mess = ["FTPtime", str(datetime.today())])
        except Exception as Err:
            print("Error Cleaning to FTP")
            ToLog("Error Cleaning to FTP, Error code = " + str(Err))
            #raise Exception
            

    @info_method_dec
    def CreateDir(self, ftp, dirname):
        try:
            ftp.cwd(dirname)
        except Exception:
            ftp.mkd(dirname)
            ftp.cwd(dirname)

#=============================================
#=============================================
#=============================================
#=============================================
#ConvertThread
class ConvertThread(threading.Thread):
    @except_method_dec
    def __init__(self, path_to_file = "some_file", path_to_dir = "some_dir"):
        super().__init__()
        self.file = path_to_file
        self.dir = path_to_dir

    @except_method_brief
    def run(self):
        ToLog(f"Convert thread started wuth {self.dir}")
        self.CreateStyles()
        conn = sqlite3.connect(self.file)
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type = 'table';")
        self.tables = [table[0] for table in cursor.fetchall()]
            
        wb = openpyxl.Workbook()
        for table in self.tables:
            wb.create_sheet(table)
            work = wb[table]

            #read column names
            cursor = conn.execute("SELECT name, type FROM pragma_table_info('" + table + "')")
            labels = [row[0] for row in cursor]

            for i in range (0, len(labels)):
                work.cell(row = 1, column = 1 + i, value = labels[i]).font = self.fonttable
                work.cell(row = 1, column = 1 + i).alignment = self.alightable
                work.cell(row = 1, column = 1 + i).border = self.bordertable
                    
                #read data
            cursor = conn.execute("SELECT * FROM " + table)
            raw_data = [row for row in cursor]
            
            for row in range (0, len(raw_data)):
                for col in range (0, len(raw_data[row])):
                    work.cell(row = 2 + row, column = 1 + col, value = raw_data[row][col]).font = self.fonttable
                    work.cell(row = 2 + row, column = 1 + col).alignment = self.alightable
                    work.cell(row = 2 + row, column = 1 + col).border = self.bordertable

            colsize = [8, 13, 9, 14, 11, 17, 27, 13, 15, 56, 75, 27, 27]
            for col in range (0, len(colsize)):
                work.column_dimensions[get_column_letter(col + 1)].width = colsize[col]
                

        wb.remove(wb['Sheet'])
        wb.save(self.dir + "\\result.xlsx")
        conn.close()

        os.remove(self.file)
        os.startfile(os.path.realpath(self.dir + "\\result.xlsx"))                     

    @except_method_brief
    def CreateStyles(self):
        self.fonttable = Font(name="Times New Roman", size=12, bold=False)
        
        self.bordertable = Border(left=Side(border_style="thin", color="FF000000"),
                             right=Side(border_style="thin", color="FF000000"),
                             top=Side(border_style="thin", color="FF000000"),
                             bottom=Side(border_style="thin", color="FF000000"))

        self.alightable = Alignment(horizontal = "center",
                               vertical = "center",
                               text_rotation = 0,
                               wrap_text = False,
                               shrink_to_fit = True,
                               indent = 0)
         
#@except_foo_dec
def Dt_to_txt(datetime):
    return str(datetime)[8:10] + "." + str(datetime)[5:7] + "." + str(datetime)[0:4]        
#=============================================
#=============================================
#=============================================
#=============================================
# scaling bitmap
@except_foo_dec
def ScaleBitmap(bitmap, size):
    image = bitmap.ConvertToImage()
    image = image.Scale(size[0], size[1], wx.IMAGE_QUALITY_HIGH)
    return wx.Image(image).ConvertToBitmap()
#=============================================
#=============================================
#=============================================
#=============================================
#Copy file foo
@except_foo_dec
def CopyFile(source, dist, buffer = 1024*1024):
    ToLog("CopyFile " + source + " to " + dist + " function started")
    with open(source, "rb") as SrcFile, open(dist, "wb") as DestFile:
        while True:
            copy_buffer = SrcFile.read(buffer)
            if not copy_buffer:
                break
            DestFile.write(copy_buffer)
    ToLog("CopyFile " + source + " to " + dist + " function finished")

#=============================================
#=============================================
#=============================================
#=============================================
# ClearOldLogs
@except_foo_dec
def ClearLogs(Dir, numfiles = int(10)):
    while len(os.listdir(Dir)) >= numfiles:
        if len(os.listdir(Dir)) < numfiles:
                break
        try:
            os.remove(os.path.abspath(FindOldest(Dir)))
            ToLog(f"foo ClearLogs, dir = {Dir}: DELETING FILE " + str(FindOldest(Dir)))
        except Exception as Err:
            ToLog(f"foo ClearLogs, dir = {Dir}: Old file with logs wasn't deleted, Error code = " + str(Err))
            #raise Exception
            break

#=============================================
#=============================================
#=============================================
#=============================================   
# DeleteOldest
@except_foo_dec
def FindOldest(Dir):
    List = os.listdir(Dir)
    fullPath = [Dir + "/{0}".format(x) for x in List]
    oldestFile = min(fullPath, key = os.path.getctime)
    return oldestFile
    
#=============================================
#=============================================
#=============================================
#=============================================
@except_foo_dec
def FindMyDir(nameDir, subDirs = None):
    if "Documents" in os.listdir(os.path.expanduser("~")):
        DocDir = os.path.expanduser("~") + "\\Documents"
    else:
        os.mkdir(os.path.expanduser("~") + "\\Documents")
        DocDir = os.path.expanduser("~") + "\\Documents"
    if nameDir not in os.listdir(DocDir):
        os.mkdir(DocDir + "\\" + nameDir)
        ToLog(nameDir + "folder was Created")

    DocDir = DocDir + "\\" + nameDir
    if isinstance (subDirs, list):
        for word in subDirs:
            if word not in os.listdir(DocDir):
                os.mkdir(DocDir + "\\" + word)
                ToLog(word + " folder was Created")
    return DocDir
    
#===============================================
#===============================================
#===============================================
#===============================================        
# Создание класса окна любой ошибки
@except_foo_brief
def SomeError(parent, title):
    wx.MessageBox(title, "Ошибка", wx.OK)


#=============================================
#=============================================
#=============================================
#=============================================
if __name__ == '__main__':
    # Определение локали!
    locale.setlocale(locale.LC_ALL, "")

    global LogDir, MonitLogDir, MyDate, threadLog
    MyDate = "22.01.2025"
    MonitOpen = False

    ToLog("\n\n" + "!" * 40)
    ToLog("Application started")

    DocDir = FindMyDir(nameDir = "Monit_SpiskiSS_Files", subDirs = ["Script_Logs", "Monitoring_Logs", "Based", "Temp"])
    LogDir = DocDir + "\\Script_Logs"
    MonitLogDir = DocDir + "\\Monitoring_Logs"
    ClearLogs(LogDir)
    ClearLogs(MonitLogDir)

    threadLog = LogThread(logdir = LogDir)
    threadLog.setDaemon(True)
    threadLog.start()

    ex = wx.App()

    HelloFrame(path_to_png = os.getcwd() + "\\images\\WritingPNG.png")
    global Frame_Osn
    Frame_Osn = Main_Frame(None, DocDir = DocDir)

    ex.MainLoop()





